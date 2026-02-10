import io
import re
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class FinancialMetric:
    metric: str
    value: float
    currency: str
    year: int | None
    context: str
    source_file: str
    page: int


FINANCIAL_PATTERNS = {
    "Revenue": [r"revenue", r"total\s+income", r"turnover", r"sales", r"income\s+from\s+operations"],
    "Gross Profit": [r"gross\s+profit"],
    "Operating Income": [r"operating\s+income", r"operating\s+profit", r"ebit"],
    "EBITDA": [r"ebitda"],
    "Net Income": [r"net\s+income", r"net\s+profit", r"profit\s+after\s+tax", r"pat"],
    "Total Assets": [r"total\s+assets"],
    "Total Liabilities": [r"total\s+liabilities"],
    "Cash Flow": [r"cash\s+flow", r"net\s+cash\s+from\s+operating\s+activities"],
    "EPS": [r"earnings\s+per\s+share", r"\beps\b"],
}

YEAR_PATTERN = re.compile(r"\b(19\d{2}|20\d{2})\b")
AMOUNT_PATTERN = re.compile(r"\(?[-+]?[$€£₹]?\s?\d[\d,]*(?:\.\d+)?\)?")


def normalize_amount(raw: str) -> float:
    """Convert text amounts like '(1,200.00)' or '₹ 3,450' to float."""
    txt = (
        raw.replace(",", "")
        .replace("$", "")
        .replace("€", "")
        .replace("£", "")
        .replace("₹", "")
        .strip()
    )
    negative = txt.startswith("(") and txt.endswith(")")
    txt = txt.strip("()")
    value = float(txt)
    return -value if negative else value


def detect_currency(text: str) -> str:
    lower = text.lower()
    if "inr" in lower or "rs." in lower or "rs " in lower or "rupee" in lower or "₹" in text:
        return "INR"
    if "usd" in lower or "$" in text:
        return "USD"
    if "eur" in lower or "€" in text:
        return "EUR"
    if "gbp" in lower or "£" in text:
        return "GBP"
    return "Unknown"


def extract_years(text: str) -> list[int]:
    return [int(y) for y in YEAR_PATTERN.findall(text)]


def assign_years_to_amounts(years: list[int], amount_count: int) -> list[int | None]:
    """
    Best-effort year mapping for annual statements:
    - If counts match, align in order.
    - If a single year exists, use it for all amounts.
    - If we have more years than amounts, keep the right-most years (common in annual columns).
    """
    if amount_count == 0:
        return []
    if not years:
        return [None] * amount_count
    if len(years) == amount_count:
        return years
    if len(years) == 1:
        return [years[0]] * amount_count
    if len(years) > amount_count:
        return years[-amount_count:]
    # fewer years than amounts and multiple years: apply latest year as fallback
    return [years[-1]] * amount_count


def extract_financial_metrics(pdf_path: Path) -> list[FinancialMetric]:
    metrics: list[FinancialMetric] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

            for line in lines:
                normalized_line = line.lower()
                for metric_name, aliases in FINANCIAL_PATTERNS.items():
                    if any(re.search(alias, normalized_line) for alias in aliases):
                        amounts = [a.replace(" ", "") for a in AMOUNT_PATTERN.findall(line) if re.search(r"\d", a)]
                        years = extract_years(line)
                        mapped_years = assign_years_to_amounts(years, len(amounts))

                        for idx, amount in enumerate(amounts):
                            try:
                                value = normalize_amount(amount)
                            except ValueError:
                                continue

                            metrics.append(
                                FinancialMetric(
                                    metric=metric_name,
                                    value=value,
                                    currency=detect_currency(line),
                                    year=mapped_years[idx] if idx < len(mapped_years) else None,
                                    context=line,
                                    source_file=pdf_path.name,
                                    page=page_number,
                                )
                            )
    return metrics


def metrics_to_dataframe(items: Iterable[FinancialMetric]) -> pd.DataFrame:
    df = pd.DataFrame([m.__dict__ for m in items])
    if df.empty:
        return pd.DataFrame(
            columns=["source_file", "page", "year", "metric", "currency", "value", "context"]
        )
    return df[["source_file", "page", "year", "metric", "currency", "value", "context"]].sort_values(
        ["source_file", "year", "metric"], na_position="last"
    )


def build_excel_report(data: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        data.to_excel(writer, index=False, sheet_name="Financial Data")

        summary = (
            data.groupby(["year", "metric", "currency"], dropna=False)["value"]
            .agg(["count", "sum", "mean"])
            .reset_index()
            .sort_values(["year", "metric"], na_position="last")
        )
        summary.to_excel(writer, index=False, sheet_name="Yearly Summary")

        pivot = (
            data.pivot_table(
                index=["metric", "currency"],
                columns="year",
                values="value",
                aggfunc="sum",
                fill_value=0,
            )
            .reset_index()
        )
        pivot.to_excel(writer, index=False, sheet_name="Metric x Year")

        workbook = writer.book
        for sheet_name in ["Financial Data", "Yearly Summary", "Metric x Year"]:
            ws = workbook[sheet_name]
            header_fill = PatternFill(start_color="0A3D62", end_color="0A3D62", fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 70)

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top")

            for col_idx, cell in enumerate(ws[1], start=1):
                header = str(cell.value or "").lower()
                if header in {"value", "sum", "mean"} or header.isdigit():
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

    output.seek(0)
    return output.read()


def main() -> None:
    st.set_page_config(page_title="PDF Financial Data Extractor", page_icon="📊", layout="wide")
    st.markdown(
        """
        <style>
            .stApp {background: linear-gradient(180deg, #f8f9ff 0%, #f0f6ff 100%);}
            .block-container {padding-top: 2rem;}
            .hero {background: white; padding: 1.2rem 1.6rem; border-radius: 14px; box-shadow: 0 10px 30px rgba(10, 61, 98, .08);}
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        '<div class="hero"><h1>📄➡️📊 PDF to Financial Excel Report</h1>'
        "Upload annual statement PDFs. The app extracts figures, attempts year alignment, and creates a styled Excel report.</div>",
        unsafe_allow_html=True,
    )

    uploaded_files = st.file_uploader("Drop PDF files here", type=["pdf"], accept_multiple_files=True)

    if not uploaded_files:
        st.info("Upload at least one PDF to begin.")
        return

    all_metrics: list[FinancialMetric] = []
    progress = st.progress(0)

    for idx, file in enumerate(uploaded_files, start=1):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(file.read())
            temp_path = Path(tmp.name)

        all_metrics.extend(extract_financial_metrics(temp_path))
        progress.progress(idx / len(uploaded_files))

    df = metrics_to_dataframe(all_metrics)

    st.subheader("Extracted Financial Rows")
    st.dataframe(df, use_container_width=True, height=420)

    if not df.empty:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Rows", f"{len(df):,}")
        c2.metric("PDF Files", f"{df['source_file'].nunique():,}")
        c3.metric("Metrics", f"{df['metric'].nunique():,}")
        c4.metric("Years", f"{df['year'].dropna().nunique():,}")

    excel_bytes = build_excel_report(df)
    st.download_button(
        label="⬇️ Download Excel Report",
        data=excel_bytes,
        file_name="financial_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
